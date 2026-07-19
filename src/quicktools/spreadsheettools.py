"""Spreadsheet utilities: reading, writing, and converting between CSV, Excel, and JSON."""
import csv
import json


def read_csv_file(path: str) -> list[dict]:
    """Read a CSV file and return its rows as a list of dictionaries (keyed by header)."""
    with open(path, "r", encoding="utf-8", newline="") as f:
        return list(csv.DictReader(f))


def write_csv_file(path: str, data: list[dict]) -> None:
    """Write a list of dictionaries to a CSV file, using the first row's keys as headers."""
    if not data:
        raise ValueError("data must contain at least one row")
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(data[0].keys()))
        writer.writeheader()
        writer.writerows(data)


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """Convert a CSV file into an Excel .xlsx file."""
    from openpyxl import Workbook
    rows = read_csv_file(csv_path)
    wb = Workbook()
    ws = wb.active
    if rows:
        ws.append(list(rows[0].keys()))
        for row in rows:
            ws.append(list(row.values()))
    wb.save(xlsx_path)


def xlsx_to_csv(xlsx_path: str, csv_path: str, sheet_name: str | None = None) -> None:
    """Convert an Excel .xlsx file (a single sheet) into a CSV file."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)


def json_to_csv(json_path: str, csv_path: str) -> None:
    """Convert a JSON file (a list of flat objects) into a CSV file."""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    write_csv_file(csv_path, data)


def csv_to_json(csv_path: str, json_path: str) -> None:
    """Convert a CSV file into a JSON file (a list of objects)."""
    data = read_csv_file(csv_path)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)


def get_xlsx_sheet_names(path: str) -> list[str]:
    """Return the list of sheet names in an Excel workbook."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    return wb.sheetnames


def merge_csv_files(paths: list[str], output_path: str) -> None:
    """Merge multiple CSV files (with the same columns) into a single CSV file."""
    all_rows = []
    for path in paths:
        all_rows.extend(read_csv_file(path))
    write_csv_file(output_path, all_rows)